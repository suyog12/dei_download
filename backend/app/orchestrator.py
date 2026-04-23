"""Orchestrator for fetching, merging, and exporting indicator data.

Flow:
    1. Resolve the requested pillars + years to a concrete indicator list
       from the catalog.
    2. Group API-available indicators by source.
    3. Fan out parallel fetches; collect DataPoint rows.
    4. Emit an xlsx workbook with one sheet per pillar, plus a Manual sheet
       listing subscription/manual sources.
"""

from __future__ import annotations

import asyncio
import io
import logging
from dataclasses import dataclass

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .catalog import CATALOG, Availability, Indicator, Pillar, by_source, indicators_for
from .sources import imf, owid, worldbank
from .sources.base import DataPoint

logger = logging.getLogger(__name__)


@dataclass
class SourceStatus:
    source: str
    requested: int  # indicators we tried
    fetched_rows: int  # DataPoint rows received
    ok: bool
    message: str = ""


@dataclass
class FetchResult:
    points: list[DataPoint]
    statuses: list[SourceStatus]


PILLAR_ORDER = [Pillar.SUPPLY, Pillar.DEMAND, Pillar.INSTITUTIONAL, Pillar.INNOVATION]
PILLAR_LABELS = {
    Pillar.SUPPLY: "Supply Conditions",
    Pillar.DEMAND: "Demand Conditions",
    Pillar.INSTITUTIONAL: "Institutional Environment",
    Pillar.INNOVATION: "Innovation and Change",
}


# Source name -> module with fetch_indicators(...).
# Only verified-working sources are registered here. Sources that were once
# API-accessible but now require manual download live in the catalog with
# availability=MANUAL and are not dispatched.
API_DISPATCH: dict[str, object] = {
    "World Bank": worldbank,
    "World Bank Global Findex": worldbank,
    "World Bank WGI": worldbank,
    "World Bank / WIPO": worldbank,
    "World Bank / UNESCO": worldbank,
    "Our World in Data": owid,
    "IMF": imf,
}


async def fetch(
    pillars: set[Pillar],
    start_year: int,
    end_year: int,
    countries: set[str] | None = None,
    indicator_keys: set[str] | None = None,
) -> FetchResult:
    """Fetch API-available indicators for the given pillars and year range.

    countries: if provided, filter returned DataPoints to these ISO-3 codes.
    indicator_keys: if provided, restrict to these catalog keys (intersected
        with pillars). This is how source-level and per-indicator toggles
        reach the fetch layer.
    """
    indicators = indicators_for(pillars, availability={Availability.API})
    if indicator_keys is not None:
        indicators = [i for i in indicators if i.key in indicator_keys]
    grouped = by_source(indicators)

    tasks = []
    task_meta: list[tuple[str, int]] = []

    for source_name, ind_list in grouped.items():
        module = API_DISPATCH.get(source_name)
        if module is None:
            logger.warning("No dispatch for source %s", source_name)
            continue
        tasks.append(module.fetch_indicators(ind_list, start_year, end_year))
        task_meta.append((source_name, len(ind_list)))

    raw_results = await asyncio.gather(*tasks, return_exceptions=True)

    all_points: list[DataPoint] = []
    statuses: list[SourceStatus] = []

    for (source_name, requested), result in zip(task_meta, raw_results):
        if isinstance(result, Exception):
            logger.exception("Source %s raised", source_name)
            statuses.append(
                SourceStatus(
                    source=source_name,
                    requested=requested,
                    fetched_rows=0,
                    ok=False,
                    message=str(result),
                )
            )
            continue
        points = result  # list[DataPoint]
        if countries is not None:
            points = [p for p in points if p.country_iso3 in countries]
        all_points.extend(points)
        # Three distinct "empty" states, each with a distinct user message:
        #   - source didn't return anything at all (likely unreachable or
        #     API rejected the request): "No rows returned from <source>"
        #   - source returned data but all of it was for countries we
        #     didn't ask for: "No matching rows for selected countries"
        #   - source returned matching rows: message is empty, ok=True
        if len(points) > 0:
            message = ""
            ok = True
        elif len(result) == 0:
            message = f"{source_name} returned no rows (endpoint may be down or indicator ID invalid)"
            ok = False
        else:
            message = "Returned rows but none for selected countries"
            ok = False
        statuses.append(
            SourceStatus(
                source=source_name,
                requested=requested,
                fetched_rows=len(points),
                ok=ok,
                message=message,
            )
        )

    return FetchResult(points=all_points, statuses=statuses)


def build_workbook(
    result: FetchResult, pillars: set[Pillar], start_year: int, end_year: int
) -> bytes:
    """Build an xlsx with one sheet per pillar + a manual-sources sheet."""
    wb = openpyxl.Workbook()
    # remove the default sheet, we'll add our own
    wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_align = Alignment(horizontal="left", vertical="center")

    # Overview sheet
    overview = wb.create_sheet("Overview")
    overview["A1"] = "Digital Evolution Index - Indicator Bundle"
    overview["A1"].font = Font(bold=True, size=14)
    overview["A3"] = "Year range"
    overview["B3"] = f"{start_year} - {end_year}"
    overview["A4"] = "Pillars selected"
    overview["B4"] = ", ".join(PILLAR_LABELS[p] for p in PILLAR_ORDER if p in pillars)
    overview["A6"] = "Source"
    overview["B6"] = "Indicators requested"
    overview["C6"] = "Rows fetched"
    overview["D6"] = "Status"
    overview["E6"] = "Notes"
    for cell in ("A6", "B6", "C6", "D6", "E6"):
        overview[cell].font = header_font
        overview[cell].fill = header_fill
        overview[cell].alignment = header_align

    for i, status in enumerate(result.statuses, start=7):
        overview.cell(row=i, column=1, value=status.source)
        overview.cell(row=i, column=2, value=status.requested)
        overview.cell(row=i, column=3, value=status.fetched_rows)
        overview.cell(row=i, column=4, value="OK" if status.ok else "Failed / empty")
        overview.cell(row=i, column=5, value=status.message)

    for col_idx, width in enumerate([34, 22, 15, 18, 50], start=1):
        overview.column_dimensions[get_column_letter(col_idx)].width = width

    # One sheet per pillar
    by_pillar: dict[Pillar, list[DataPoint]] = {p: [] for p in PILLAR_ORDER}
    for pt in result.points:
        try:
            p = Pillar(pt.pillar)
        except ValueError:
            continue
        by_pillar[p].append(pt)

    columns = [
        ("Country", "country_name", 28),
        ("ISO3", "country_iso3", 8),
        ("Year", "year", 8),
        ("Indicator", "indicator_name", 60),
        ("Value", "value", 14),
        ("Unit", "unit", 14),
        ("Component", "component", 30),
        ("Source", "source", 24),
        ("Indicator Key", "indicator_key", 26),
    ]

    for pillar in PILLAR_ORDER:
        if pillar not in pillars:
            continue
        sheet = wb.create_sheet(PILLAR_LABELS[pillar])

        for col_idx, (label, _attr, width) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            sheet.column_dimensions[get_column_letter(col_idx)].width = width

        rows = by_pillar.get(pillar, [])
        # Sort for readability: country, indicator, year
        rows.sort(key=lambda r: (r.country_name, r.indicator_name, r.year))

        for r_idx, pt in enumerate(rows, start=2):
            for c_idx, (_label, attr, _w) in enumerate(columns, start=1):
                sheet.cell(row=r_idx, column=c_idx, value=getattr(pt, attr))

        sheet.freeze_panes = "A2"

    # Manual / subscription sheet
    manual_inds = [i for i in CATALOG if i.availability != Availability.API and i.pillar in pillars]
    if manual_inds:
        manual = wb.create_sheet("Manual & Subscription Sources")
        headers = ["Source", "Indicator", "Pillar", "Availability", "URL", "Notes"]
        for col_idx, h in enumerate(headers, start=1):
            cell = manual.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        widths = [22, 48, 16, 16, 48, 60]
        for i, w in enumerate(widths, start=1):
            manual.column_dimensions[get_column_letter(i)].width = w

        for r_idx, ind in enumerate(manual_inds, start=2):
            manual.cell(row=r_idx, column=1, value=ind.source)
            manual.cell(row=r_idx, column=2, value=ind.name)
            manual.cell(row=r_idx, column=3, value=PILLAR_LABELS[ind.pillar])
            manual.cell(row=r_idx, column=4, value=ind.availability.value)
            manual.cell(row=r_idx, column=5, value=ind.manual_url)
            manual.cell(row=r_idx, column=6, value=ind.notes)

        manual.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
