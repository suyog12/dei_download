"""Anthropic Economic Index (AEI) fetcher.

Fetches country-level AI adoption metrics from the Hugging Face dataset
published at https://huggingface.co/datasets/Anthropic/EconomicIndex.

AEI data is NOT a country-by-year time series like WB/IMF. It's a snapshot
of AI adoption during a specific data collection window (2025-08-04 through
2025-08-11 for the release used here). For that reason it lives in its own
panel in the UI rather than being mixed with DEI pillars.

File used:
    release_2025_09_15/data/output/aei_enriched_claude_ai_2025-08-04_to_2025-08-11.csv

Schema (long format, one row per {country, variable} pair):
    geo_id, geography, date_start, date_end, platform_and_product, facet,
    level, variable, cluster_name, value

We filter to rows where:
  - geography == "country"
  - facet == "country"
  - variable in COUNTRY_VARIABLES (defined below)

CC-BY licensed (data) / MIT (code).
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass

import httpx

logger = logging.getLogger(__name__)

# Pinned to the 2025-09-15 release which has verified geographic data.
# Newer releases (2026-01-15, 2026-03-24) are focused on different axes.
AEI_CSV_URL = (
    "https://huggingface.co/datasets/Anthropic/EconomicIndex/resolve/main/"
    "release_2025_09_15/data/output/"
    "aei_enriched_claude_ai_2025-08-04_to_2025-08-11.csv"
)
UA = "DEI-Downloader/1.0 (+research tool)"

# The variables worth surfacing in the country panel. Anything else is
# either operational (counts raw), secondary (per-capita without index),
# or not meaningful at country scope.
COUNTRY_VARIABLES: dict[str, str] = {
    "usage_per_capita_index": "AI Usage per Capita Index (1.0 = proportional to population)",
    "usage_pct": "Share of global Claude.ai conversations (%)",
    "automation_pct": "Share of conversations showing automation patterns (%)",
    "augmentation_pct": "Share of conversations showing augmentation patterns (%)",
    "usage_tier": "Usage adoption tier (0-4)",
}


@dataclass
class AEICountryRow:
    """One country's AEI metrics. Not tied to a year — AEI is a snapshot."""

    country_iso3: str
    country_name: str
    date_start: str
    date_end: str
    variable_key: str  # machine-readable, e.g. "usage_pct"
    variable_label: str  # human-readable, for column headers
    value: float | None


# Simple ISO-2 -> ISO-3 map for countries AEI commonly includes. The AEI raw
# file is keyed by ISO-2 (e.g. "US"); some enrichment steps produce ISO-3
# already but we defend either way.
_ISO2_TO_ISO3 = {
    "US": "USA", "GB": "GBR", "CA": "CAN", "AU": "AUS", "NZ": "NZL",
    "DE": "DEU", "FR": "FRA", "ES": "ESP", "IT": "ITA", "NL": "NLD",
    "BE": "BEL", "SE": "SWE", "NO": "NOR", "DK": "DNK", "FI": "FIN",
    "IE": "IRL", "PT": "PRT", "PL": "POL", "CZ": "CZE", "AT": "AUT",
    "CH": "CHE", "GR": "GRC", "RO": "ROU", "HU": "HUN", "IL": "ISR",
    "JP": "JPN", "KR": "KOR", "TW": "TWN", "SG": "SGP", "HK": "HKG",
    "IN": "IND", "ID": "IDN", "MY": "MYS", "PH": "PHL", "TH": "THA",
    "VN": "VNM", "BR": "BRA", "MX": "MEX", "AR": "ARG", "CL": "CHL",
    "CO": "COL", "PE": "PER", "ZA": "ZAF", "EG": "EGY", "NG": "NGA",
    "KE": "KEN", "TR": "TUR", "SA": "SAU", "AE": "ARE", "QA": "QAT",
    "CN": "CHN", "RU": "RUS", "UA": "UKR", "BG": "BGR", "HR": "HRV",
    "SK": "SVK", "SI": "SVN", "LT": "LTU", "LV": "LVA", "EE": "EST",
}


async def fetch_country_rows(
    variables: list[str] | None = None,
) -> list[AEICountryRow]:
    """Fetch and parse the AEI country-level CSV.

    Args:
        variables: If provided, return only rows for these variable keys.
            Default is all of COUNTRY_VARIABLES.
    """
    target_vars = set(variables) if variables else set(COUNTRY_VARIABLES.keys())

    async with httpx.AsyncClient(headers={"User-Agent": UA}) as client:
        try:
            resp = await client.get(AEI_CSV_URL, timeout=60.0, follow_redirects=True)
        except httpx.RequestError as exc:
            logger.warning("AEI request error: %s", exc)
            return []

        if resp.status_code != 200:
            body_preview = resp.text[:200].replace("\n", " ")
            logger.warning(
                "AEI returned HTTP %s: %s", resp.status_code, body_preview
            )
            return []

    text = resp.text
    # Hugging Face occasionally returns an HTML LFS redirect page if the
    # dataset uses xet/LFS and follow_redirects missed a hop. Bail defensively.
    if text.lstrip().startswith("<") or "," not in text[:200]:
        logger.warning("AEI response did not look like CSV (len=%d)", len(text))
        return []

    reader = csv.DictReader(io.StringIO(text))
    fieldnames = reader.fieldnames or []
    required = {"geo_id", "geography", "facet", "variable", "cluster_name", "value",
                "date_start", "date_end"}
    missing = required - set(fieldnames)
    if missing:
        logger.warning("AEI CSV missing expected columns: %s", missing)
        return []

    out: list[AEICountryRow] = []
    for row in reader:
        # We only want country-level rows for the country facet.
        if row.get("geography") != "country" or row.get("facet") != "country":
            continue
        var = (row.get("variable") or "").strip()
        if var not in target_vars:
            continue

        geo = (row.get("geo_id") or "").strip().upper()
        if len(geo) == 2:
            iso3 = _ISO2_TO_ISO3.get(geo, "")
        elif len(geo) == 3:
            iso3 = geo
        else:
            continue
        if not iso3:
            continue

        # cluster_name is the country itself for country-faceted rows, but
        # can be "not_classified" which we skip.
        cluster = (row.get("cluster_name") or "").strip()
        if cluster in ("not_classified", "none", ""):
            continue

        raw = (row.get("value") or "").strip()
        try:
            value = float(raw) if raw else None
        except ValueError:
            value = None

        out.append(
            AEICountryRow(
                country_iso3=iso3,
                country_name=cluster,
                date_start=row.get("date_start") or "",
                date_end=row.get("date_end") or "",
                variable_key=var,
                variable_label=COUNTRY_VARIABLES[var],
                value=value,
            )
        )

    return out


def build_workbook(rows: list[AEICountryRow]) -> bytes:
    """Build an xlsx for AEI country data in a pivoted, human-friendly form.

    One row per country with columns for each of the COUNTRY_VARIABLES plus
    country metadata. Returns the xlsx file as bytes.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    ws = wb.active
    ws.title = "AEI Country Index"

    # Pivot long -> wide: {(iso3, var): value}
    pivot: dict[tuple[str, str], float | None] = {}
    countries: dict[str, str] = {}  # iso3 -> human-readable name
    date_range = ""
    for r in rows:
        pivot[(r.country_iso3, r.variable_key)] = r.value
        countries[r.country_iso3] = r.country_name
        if not date_range:
            date_range = f"{r.date_start} to {r.date_end}"

    var_order = list(COUNTRY_VARIABLES.keys())
    header = ["Country", "ISO3"] + [COUNTRY_VARIABLES[v] for v in var_order]
    ws.append(["Anthropic Economic Index — Country panel"])
    ws.append([f"Data collection window: {date_range or 'unknown'}"])
    ws.append([])
    ws.append(header)
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.cell(row=4, column=1).font = Font(bold=True)
    for col_idx in range(1, len(header) + 1):
        ws.cell(row=4, column=col_idx).font = Font(bold=True)

    # Sort by usage_per_capita_index descending when available, else by name
    def _sort_key(iso3: str) -> tuple[int, float]:
        val = pivot.get((iso3, "usage_per_capita_index"))
        if val is None:
            return (1, 0.0)
        return (0, -val)

    for iso3 in sorted(countries.keys(), key=_sort_key):
        row = [countries[iso3], iso3]
        for var in var_order:
            row.append(pivot.get((iso3, var)))
        ws.append(row)

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 8
    for i in range(len(var_order)):
        col = chr(ord("C") + i)
        ws.column_dimensions[col].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
